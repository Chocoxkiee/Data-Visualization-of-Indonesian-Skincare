from openpyxl import load_workbook
import matplotlib.pyplot as plt

#Load the workbook and select the first sheet
workbook = load_workbook(filename="Indonesian Skincare Dataset.xlsx")
sheet = workbook.active

#Extract column headers
headers = [cell.value for cell in sheet[1]]

#Find indexes for 'Brand' and 'Rating' columns
brand_col = headers.index('Brand') + 1
rating_col = headers.index('Rating') + 1

#Extract data (excluding header)
brands = []  
ratings = []  

for row in sheet.iter_rows(min_row=2, max_row=51, values_only=True):
    brand_name = row[brand_col - 1]  
    rating = row[rating_col - 1]     
    if brand_name and rating:
        brands.append(brand_name)  
        ratings.append(float(rating))  

#Plotting
plt.figure(figsize=(10, 6))
plt.barh(brands, ratings, color='lightgreen')
plt.title("Product Ratings - Horizontal Bar")
plt.xlabel("Rating")
plt.ylabel("Brand")
plt.tight_layout()
plt.show()

