from openpyxl import load_workbook
import matplotlib.pyplot as plt
from collections import defaultdict
import numpy as np

# Load workbook and select active sheet
workbook = load_workbook(filename="Indonesian Skincare Dataset.xlsx")
sheet = workbook.active

# Get column indexes
headers = [cell.value for cell in sheet[1]]
category_col = headers.index('Category') + 1
brand_col = headers.index('Brand') + 1
rating_col = headers.index('Rating') + 1

# Prepare data
data = defaultdict(lambda: defaultdict(list))  # category -> brand -> list of ratings

for row in sheet.iter_rows(min_row=2, max_row=31, values_only=True):  # Limit to 100 rows
    category = row[category_col - 1]
    brand = row[brand_col - 1]
    rating = row[rating_col - 1]
    if category and brand and rating:
        data[category][brand].append(float(rating))

# Compute average ratings
categories = list(data.keys())
brands_set = set()
for brand_dict in data.values():
    brands_set.update(brand_dict.keys())
brands = sorted(list(brands_set))

avg_ratings = {brand: [] for brand in brands}

for category in categories:
    brand_dict = data[category]
    for brand in brands:
        if brand in brand_dict:
            avg_rating = sum(brand_dict[brand]) / len(brand_dict[brand])
        else:
            avg_rating = 0
        avg_ratings[brand].append(avg_rating)

# Plotting
x = np.arange(len(categories))
width = 0.8 / len(brands)  # adjust bar width depending on number of brands

plt.figure(figsize=(14, 6))

for i, brand in enumerate(brands):
    plt.bar(x + i * width, avg_ratings[brand], width=width, label=brand)

plt.xlabel("Category")
plt.ylabel("Average Rating")
plt.title("Average Product Ratings by Category and Brand")
plt.xticks(x + width * len(brands) / 2, categories, rotation=45, ha='right')
plt.legend(title="Brand", bbox_to_anchor=(1.05, 1), loc='upper left')
plt.tight_layout()
plt.show()
