from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Load the workbook and select the first sheet
workbook = load_workbook(filename="Indonesian Skincare Dataset.xlsx")
sheet = workbook.active

# Extract column headers
headers = [cell.value for cell in sheet[1]]

# Find indexes for 'Category', 'Brand', 'Rating', and 'Total Reviewers' columns
category_col = headers.index('Category') + 1
brand_col = headers.index('Brand') + 1
rating_col = headers.index('Rating') + 1
reviewer_col = headers.index('Total Reviewers') + 1

# Data containers for Sunscreen category
sunscreen_brands = []
sunscreen_ratings = []
sunscreen_reviewers = []

# Loop through the rows
for row in sheet.iter_rows(min_row=2, max_row=16, values_only=True):
    category = row[category_col - 1]
    brand = row[brand_col - 1]
    rating = row[rating_col - 1]
    reviewer = row[reviewer_col - 1]
    
    if category == 'Sunscreen' and rating and reviewer:
        sunscreen_brands.append(brand)
        sunscreen_ratings.append(float(rating))
        sunscreen_reviewers.append(int(reviewer))

# Pie Chart using Total Reviewers for Sunscreen products
plt.figure(figsize=(6, 6))
plt.pie(sunscreen_reviewers, labels=sunscreen_brands, autopct='%1.1f%%', startangle=140)
plt.title("Sunscreen Product Proportions - Based on Total Reviewers")
plt.tight_layout()
plt.show()
