from openpyxl import load_workbook
import matplotlib.pyplot as plt

#Load the workbook and select the first sheet
workbook = load_workbook(filename="Indonesian Skincare Dataset.xlsx")
sheet = workbook.active

#Extract column headers
headers = [cell.value for cell in sheet[1]]

#Find indexes for 'Brand' and 'Rating' columns
brand_col = headers.index('Brand') + 1
rating_col = headers.index('Rating') + 1

#Extract data (excluding header)
brands = []  
ratings = []  

for row in sheet.iter_rows(min_row=2, max_row=51, values_only=True):
    brand_name = row[brand_col - 1]  
    rating = row[rating_col - 1]     
    if brand_name and rating:
        brands.append(brand_name)  
        ratings.append(float(rating))  

# Get top 10 based on rating
top_10 = sorted(zip(brands, ratings), key=lambda x: x[1], reverse=True)[:10]
top_10_brands, top_10_ratings = zip(*top_10)

plt.figure(figsize=(6, 6))
plt.pie(top_10_ratings, labels=top_10_brands, autopct='%1.1f%%', startangle=140)
plt.title("Top 10 Brand Ratings - Pie Chart")
plt.tight_layout()
plt.show()


