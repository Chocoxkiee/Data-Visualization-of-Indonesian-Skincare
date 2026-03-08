from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Load the workbook and select the first sheet
workbook = load_workbook(filename="Indonesian Skincare Dataset.xlsx")
sheet = workbook.active

# Extract column headers
headers = [cell.value for cell in sheet[1]]

# Find indexes for 'Category', 'Brand', and 'Rating' columns
category_col = headers.index('Category') + 1
brand_col = headers.index('Brand') + 1
rating_col = headers.index('Rating') + 1

# Extract data
categories = []
brands = []
ratings = []

for row in sheet.iter_rows(min_row=2, max_row=101, values_only=True):  # Limit to first 100 rows
    category = row[category_col - 1]
    brand = row[brand_col - 1]
    rating = row[rating_col - 1]
    if category and brand and rating:
        categories.append(category)
        brands.append(brand)
        ratings.append(float(rating))

# Unique categories for color grouping
unique_categories = list(set(categories))
colors = plt.cm.tab10.colors  # color palette
category_color_map = {cat: colors[i % len(colors)] for i, cat in enumerate(unique_categories)}

# Plotting
plt.figure(figsize=(12, 6))

for cat in unique_categories:
    xs = [brands[i] for i in range(len(brands)) if categories[i] == cat]
    ys = [ratings[i] for i in range(len(ratings)) if categories[i] == cat]
    plt.scatter(xs, ys, label=cat, color=category_color_map[cat])

plt.title("Scatter Plot of Brand Ratings by Category")
plt.xlabel("Brand")
plt.ylabel("Rating")
plt.xticks(rotation=45, ha='right')
plt.legend(title="Category")
plt.tight_layout()
plt.show()
