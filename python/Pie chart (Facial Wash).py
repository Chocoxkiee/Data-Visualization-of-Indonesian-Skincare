from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Load the workbook and select the first sheet
workbook = load_workbook(filename="Indonesian Skincare Dataset.xlsx")
sheet = workbook.active

# Extract column headers
headers = [cell.value for cell in sheet[1]]

# Find indexes for 'Category', 'Brand', and 'Rating' columns
category_col = headers.index('Category') + 1
brand_col = headers.index('Brand') + 1
rating_col = headers.index('Rating') + 1

# Extract data (only facial wash category)
ratings = []
brands = []

for row in sheet.iter_rows(min_row=2, max_row=101, values_only=True):
    category = row[category_col - 1]
    brand = row[brand_col - 1]
    rating = row[rating_col - 1]
    if category == 'Facial Wash' and rating:
        ratings.append(float(rating))
        brands.append(brand)

# Create a pie chart for the ratings distribution
plt.figure(figsize=(6, 6))
plt.pie(ratings, labels=brands, autopct='%1.1f%%', startangle=140, colors=plt.cm.Paired.colors)
plt.title("Facial Wash Product Ratings Distribution")
plt.tight_layout()
plt.show()
